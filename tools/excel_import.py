"""엑셀 마이그레이션 도구 — bokju-crm Phase 2.

기존 Google Sheets 다운로드 .xlsx 를 읽어 patients/consultations 로 적재.
시트별 스키마 자동 감지 (A=구식 / B=중간 / C=신식+요일) + outlier 보정.

사용법:
    python tools/excel_import.py "uploads/상담내역 종합.xlsx" --sheet 26.5
        → dry-run 리포트만 출력. DB 변경 없음.
    python tools/excel_import.py ... --apply
        → 실제 적재. 직전 자동 백업.
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime, date
from pathlib import Path

# bokju-crm 루트를 import 경로에 추가
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

import models  # noqa: E402
from config import (  # noqa: E402
    SIGUNGU_INDEX, SIDO_LIST, COUNSELORS,
    REFERRAL_SOURCE_GROUPS, REFERRAL_TYPES,
    ADMISSION_STATUSES, CURRENT_LOCATION_TYPES,
    CONSULT_CHANNELS, DISEASES_GROUPS,
)


# ─────────────────────────────────────────────────────────
# 시트 스키마 감지
# ─────────────────────────────────────────────────────────

# 표준 헤더 (스키마별 핵심 컬럼)
SCHEMA_C_HEADERS = ("상담일자", "요일", "환자이름", "성별", "나이",
                    "대상환자군", "대상질환", "병명")
SCHEMA_B_HEADERS = ("상담일자", "환자이름", "성별", "나이",
                    "대상환자군", "대상질환", "병명")
SCHEMA_A_HEADERS = ("상담일자", "환자이름", "성별", "나이", "병명",
                    "입원 목적")


def detect_schema(rows_1to5):
    """행 1~5를 보고 (schema, header_row_idx) 반환."""
    for header_idx in (2, 3):  # 행 2 또는 3
        if len(rows_1to5) < header_idx:
            continue
        row = [str(v).strip() if v is not None else "" for v in rows_1to5[header_idx - 1]]
        joined = " | ".join(row)
        if all(h in joined for h in SCHEMA_C_HEADERS):
            return ("C", header_idx)
        if all(h in joined for h in SCHEMA_B_HEADERS) and "요일" not in joined:
            return ("B", header_idx)
        if all(h in joined for h in SCHEMA_A_HEADERS) and "대상환자군" not in joined:
            return ("A", header_idx)
    return (None, None)


def header_index_map(header_row):
    """헤더 행 → {정규화이름: 컬럼 인덱스(0-based)}.
    중복 헤더는 처음 인덱스만 채택. 좌우 공백·줄바꿈 정리."""
    m = {}
    for i, v in enumerate(header_row):
        if v is None:
            continue
        key = str(v).strip().replace("\n", "")
        if key and key not in m:
            m[key] = i
    return m


# ─────────────────────────────────────────────────────────
# 값 정규화
# ─────────────────────────────────────────────────────────

def norm_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def parse_date(v):
    """엑셀 cell → 'YYYY-MM-DD' (실패 시 None)."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # '25.12.03', '26.04.02', '2026.03.06', '25.4.1'
    m = re.match(r"^(\d{2,4})[.\-/](\d{1,2})[.\-/](\d{1,2})", s)
    if m:
        y, mo, d = m.group(1), m.group(2), m.group(3)
        if len(y) == 2:
            y = "20" + y
        try:
            return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
        except ValueError:
            return None
    # 'YYYY-MM-DD HH:MM:SS'
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def parse_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def normalize_phone(v):
    if v is None:
        return None
    digits = re.sub(r"\D", "", str(v))
    if len(digits) == 11 and digits.startswith("010"):
        return f"010-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 10:  # 02, 지역번호
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    if not digits:
        return None
    return str(v).strip()


def normalize_gender(v):
    s = norm_str(v)
    if not s:
        return None
    s_up = s.upper()
    if s_up in ("M", "남", "남성"):
        return "M"
    if s_up in ("F", "여", "여성"):
        return "F"
    return "U"


# 시도 약칭/구표기 → SIDO_LIST 풀네임 (사용자 결정: 풀네임으로 통일)
SIDO_NORMALIZE = {
    # 약칭
    "서울": "서울특별시", "부산": "부산광역시", "대구": "대구광역시",
    "인천": "인천광역시", "광주": "광주광역시", "대전": "대전광역시",
    "울산": "울산광역시", "세종": "세종특별자치시",
    "경기": "경기도", "강원": "강원특별자치도",
    "충북": "충청북도", "충남": "충청남도",
    "전북": "전북특별자치도", "전남": "전라남도",
    "경북": "경상북도", "경남": "경상남도",
    "제주": "제주특별자치도",
    # 풀네임 변형(구표기) — 모두 SIDO_LIST 표준으로
    "강원도": "강원특별자치도", "전라북도": "전북특별자치도",
    "제주도": "제주특별자치도", "세종시": "세종특별자치시",
}
# SIDO_LIST 표준값은 그대로 통과
for _full in SIDO_LIST:
    SIDO_NORMALIZE.setdefault(_full, _full)

# 복합값 첫 토큰 우선 (예: "서울 경기" → "서울특별시")
COMPOUND_SIDO_RE = re.compile(r"\s+|/|,")


def normalize_sido(v, sigungu_hint=None):
    """시도 정규화 → SIDO_LIST 풀네임. sigungu_hint(시군구) 있으면 그쪽으로 역추론 우선."""
    if sigungu_hint:
        sg = norm_str(sigungu_hint)
        if sg:
            for raw_sigungu, sidos in SIGUNGU_INDEX.items():
                if raw_sigungu == sg or raw_sigungu.startswith(sg) or sg.startswith(raw_sigungu):
                    if sidos:
                        return sidos[0]  # SIGUNGU_INDEX 값은 이미 풀네임
    s = norm_str(v)
    if not s:
        return None
    # "서울 경기" → 서울 (사용자 룰: 첫 시도 채택)
    tokens = [t for t in COMPOUND_SIDO_RE.split(s) if t]
    for t in tokens:
        if t in SIDO_NORMALIZE:
            return SIDO_NORMALIZE[t]
    # 부분 매칭 (예: "경상" → "경상북도" 첫 매칭)
    for full in SIDO_LIST:
        if s in full or full.startswith(s):
            return full
    return s  # unknown은 원문 유지 (리포트에 등장)


def normalize_sigungu(v):
    s = norm_str(v)
    if not s:
        return None
    # 직접 매칭
    if s in SIGUNGU_INDEX:
        return s
    # "안양", "수원" 등이 마스터엔 "안양시", "수원시" 형태일 수 있어 prefix 시도
    for sg in SIGUNGU_INDEX.keys():
        if sg.startswith(s) or s.startswith(sg):
            return sg
    return s


# 현재 거처 매핑: 엑셀 표기 → CURRENT_LOCATION_TYPES (집/입원중/입소중)
# 오타·표기 흔들림 모두 정규화 (사용자 결정: 급성기병운 등 → 급성기병원 = 입원중)
LOCATION_TYPE_MAP = {
    "집": "집", "자택": "집", "본인집": "집", "본인 집": "집",
    "급성기병원": "입원중", "병원": "입원중", "재활병원": "입원중",
    "요양병원": "입원중", "회복기병원": "입원중",
    # 급성기병원 오타들
    "급성기병운": "입원중", "금성기병원": "입원중",
    "급상기병원": "입원중", "급성기": "입원중",
    "요양원": "입소중", "요양원(시설)": "입소중", "시설": "입소중",
}


def normalize_location_type(v):
    s = norm_str(v)
    if not s:
        return None
    if s in LOCATION_TYPE_MAP:
        return LOCATION_TYPE_MAP[s]
    return s  # 원문 유지 (보고서에서 발견하면 매핑 추가)


# 상담방법 매핑
CONSULT_CHANNEL_MAP = {
    "전화": "전화상담", "전화상담": "전화상담",
    "방문": "내원상담", "방문상담": "내원상담", "내원": "내원상담", "내원상담": "내원상담",
}


def normalize_consult_channel(v):
    s = norm_str(v)
    if not s:
        return None
    return CONSULT_CHANNEL_MAP.get(s, s)


# 입원여부 매핑 — 상태 4종 개편(2026-05): 입원예정·입원확정은 입원보류로 통합
ADMISSION_STATUS_MAP = {
    "상담완료": "상담완료",
    "입원예정": "입원보류",
    "입원확정": "입원보류",
    "입원보류": "입원보류", "보류": "입원보류",
    "입원완료": "입원완료", "입원": "입원완료",
    "퇴원완료": "퇴원완료", "퇴원": "퇴원완료",
    "입원취소": "입원취소", "취소": "입원취소",
}


def normalize_admission_status(v):
    s = norm_str(v)
    if not s:
        return None
    return ADMISSION_STATUS_MAP.get(s, s)


# 입원목적 오타 정규화 (사용자 결정: 회복지재활 등 → 회복기재활,
#   다재내성균/다제내셩균 → 다제내성균(원문 유지·별도 통계))
ADMISSION_PURPOSE_MAP = {
    "회복지재활": "회복기재활", "회보기재활": "회복기재활",
    "비회복기": "비회복기재활",
    "다재내성균": "다제내성균", "다제내셩균": "다제내성균",
    "요양병원": "요양",
}


def normalize_admission_purpose(v):
    s = norm_str(v)
    if not s:
        return None
    return ADMISSION_PURPOSE_MAP.get(s, s)


# 유입경로(병원정보확인) → REFERRAL_TYPES (온라인/소개/기타)
# 엑셀 값들의 그룹 분류
REFERRAL_TYPE_VALUE_MAP = {
    # 온라인 (대분류 자체, 또는 세부값)
    "온라인검색": "온라인", "온라인": "온라인",
    "카페": "온라인", "검색": "온라인", "검색(블로그)": "온라인",
    "유튜브": "온라인", "SNS": "온라인", "홈페이지": "온라인",
    # 소개
    "소개": "소개", "직원추천": "소개", "지인소개": "소개",
    "기관소개": "소개", "가족회의": "소개",
    # 기타
    "기타": "기타", "지역민": "기타",
}


def normalize_referral_type(v):
    s = norm_str(v)
    if not s:
        return None
    return REFERRAL_TYPE_VALUE_MAP.get(s, s)


# ─────────────────────────────────────────────────────────
# 병명 → 4그룹 매칭
# ─────────────────────────────────────────────────────────

# 키워드 → 그룹 매핑 (포함 매칭)
DISEASE_KEYWORDS = {
    "중추신경계": ["뇌경색", "뇌출혈", "뇌손상", "뇌성마비", "척수손상", "마비", "편마비", "사지마비"],
    "근골격계": ["골절", "고관절", "대퇴", "골반", "절단", "척추협착", "추간판"],
    "비사용증후군": [
        "폐렴", "폐질환", "심장", "신생물", "다제내성", "패혈증", "농양", "균",
        "COPD", "신부전", "폐수종", "백혈병", "CRE", "VRE",
        "동정맥루", "복부대동맥류", "급성복막염", "장폐색",
        "노환", "연하곤란", "하지위약", "전신쇠약", "보행장애",
        "신장", "폐기능",
    ],
    "기저질환": ["당뇨", "고혈압", "파킨슨", "치매", "암", "이상행동"],
}

# 키워드 → DISEASES_GROUPS 멤버 라벨 (구체 라벨 우선 매핑)
# 매칭되면 by_disease Top 차트에도 정확히 노출. 매칭 안 되면 group명 폴백.
KEYWORD_TO_LABEL = {
    # 중추신경계
    "뇌경색": "뇌경색", "뇌출혈": "뇌출혈", "척수손상": "척수손상",
    "뇌성마비": "뇌성마비",
    "편마비": "마비", "사지마비": "마비", "마비": "마비",
    "뇌손상": None,  # 그룹만 표시(중추신경계). 뇌출혈/뇌경색 어느쪽인지 불명
    # 근골격계
    "고관절": "고관절 골절", "대퇴": "대퇴부 골절", "골반": "골반 골절",
    "절단": "하지 부위 절단",
    # 비사용증후군
    "폐렴": "폐질환", "폐질환": "폐질환", "폐수종": "폐질환",
    "심장": "심장질환",
    "신생물": "신생물", "백혈병": "신생물",
    # 기저질환
    "당뇨": "당뇨", "고혈압": "고혈압", "파킨슨": "파킨슨",
    "치매": "치매", "암": "암", "이상행동": "이상행동",
}


def infer_disease_groups(target_group_text, disease_text):
    """대상환자군 + 병명 텍스트 → (groups_set, matched_keywords, raw)."""
    groups = set()
    matched = []
    raw = norm_str(disease_text) or ""

    tg = norm_str(target_group_text)
    if tg:
        if tg in ("중추신경계",):
            groups.add("중추신경계")
        elif tg in ("근골격계",):
            groups.add("근골격계")
        elif tg in ("그 외", "그외", "비사용증후군", "비사용 증후군"):
            groups.add("비사용증후군")

    for grp, kws in DISEASE_KEYWORDS.items():
        for kw in kws:
            if kw in raw:
                groups.add(grp)
                matched.append((grp, kw))
                break

    return groups, matched, raw


def derive_diseases_field(target_group_text, disease_text):
    """diseases JSON 필드용 라벨 리스트 생성.
    - 키워드 매칭되어 specific 라벨이 있으면 그 라벨 사용 (예: 뇌경색)
    - 매칭은 됐지만 specific 라벨이 None이면 그룹명 사용 (예: 뇌손상→중추신경계)
    - 매칭 안 됐고 target_group으로 그룹만 알면 그룹명 사용 (예: 간농양→비사용증후군)
    """
    groups, matched, raw = infer_disease_groups(target_group_text, disease_text)
    out = []
    seen_groups = set()

    for grp, kw in matched:
        label = KEYWORD_TO_LABEL.get(kw)
        if label and label not in out:
            out.append(label)
            seen_groups.add(grp)
        elif kw in KEYWORD_TO_LABEL and label is None:
            # 키워드는 있는데 specific 라벨이 None이면 그룹명 추가 (중복 방지)
            if grp not in seen_groups and grp not in out:
                out.append(grp)
                seen_groups.add(grp)

    # target_group만으로 그룹 추론된 경우 (specific 키워드 매칭 없음)
    for grp in groups:
        if grp not in seen_groups and grp not in out:
            out.append(grp)
            seen_groups.add(grp)

    return out


# ─────────────────────────────────────────────────────────
# Schema C 파서 (26.5 등)
# ─────────────────────────────────────────────────────────

def parse_schema_c_row(headers_idx, row):
    """행 → 정규화 dict. 매핑 못 한 값은 별도 collect."""
    def cell(name):
        i = headers_idx.get(name)
        if i is None or i >= len(row):
            return None
        return row[i]

    d = {}
    issues = []

    # ── 환자 ──
    name = norm_str(cell("환자이름"))
    if not name:
        return None, ["환자이름 누락"]
    d["patient_name"] = name
    d["gender"] = normalize_gender(cell("성별"))
    d["patient_age"] = parse_int(cell("나이"))

    # 보호자
    d["guardian_name"] = norm_str(cell("이름"))
    d["guardian_relation"] = norm_str(cell("관계"))
    d["guardian_phone"] = normalize_phone(cell("연락처"))

    # 거주지: '연고지1(도)/연고지2(지역)'를 환자 거주지로
    sigungu_raw = cell("연고지2(지역)")
    sigungu_alt = cell("현 거처(지역)")
    sigungu = normalize_sigungu(sigungu_raw or sigungu_alt)
    sido_raw = cell("연고지1(도)") or cell("현 거처(도)")
    sido = normalize_sido(sido_raw, sigungu_hint=sigungu)
    d["residence_sido"] = sido
    d["residence_sigungu"] = sigungu

    # ── 상담 ──
    d["consult_date"] = parse_date(cell("상담일자"))
    if not d["consult_date"]:
        issues.append(f"상담일자 파싱 실패: {cell('상담일자')!r}")

    d["counselor"] = norm_str(cell("상담자"))
    if d["counselor"] and d["counselor"] not in COUNSELORS:
        issues.append(f"미등록 상담자: {d['counselor']}")

    d["consult_channel"] = normalize_consult_channel(cell("상담방법"))

    # 거처/모병원
    loc_raw = norm_str(cell("현재 거처"))
    d["current_location_type"] = normalize_location_type(loc_raw)
    hosp = norm_str(cell("병원이름"))
    d["source_hospital"] = hosp
    if d["current_location_type"] in ("입원중", "입소중") and hosp:
        d["current_location_name"] = hosp
    if loc_raw and not d["current_location_type"]:
        issues.append(f"미매핑 거처: {loc_raw}")

    # 유입경로
    rtype_raw = norm_str(cell("병원정보확인"))
    rtype = normalize_referral_type(rtype_raw)
    if rtype:
        d["referral_source_type"] = [rtype]
    rdetail = norm_str(cell("세부 경로"))
    if rdetail:
        d["referral_source_detail"] = [rdetail]
    if rtype_raw and rtype not in REFERRAL_TYPES:
        issues.append(f"미매핑 유입경로: {rtype_raw}")
    d["referrer_person"] = norm_str(cell("추천인"))

    # 입원목적/병명
    d["admission_purpose"] = normalize_admission_purpose(cell("입원 목적"))
    target_group = norm_str(cell("대상환자군"))
    target_disease = norm_str(cell("대상질환"))
    disease_name = norm_str(cell("병명"))
    diseases_list = derive_diseases_field(target_group, disease_name)
    if diseases_list:
        d["diseases"] = diseases_list
    # 원문 보존
    parts = [p for p in (target_disease, disease_name) if p]
    if parts:
        d["disease_detail"] = " / ".join(parts)
    if disease_name and not diseases_list:
        issues.append(f"병명 그룹 미매칭: {disease_name}")

    # 진행 상태
    d["admission_status"] = normalize_admission_status(cell("입원여부"))

    # 신규 컬럼
    iso_admit = parse_date(cell("입원일 / 비고")) or parse_date(cell("입원일/비고"))
    d["actual_admission_date"] = iso_admit
    d["recontact_memo"] = norm_str(cell("재 접촉 관리 현황")) or norm_str(cell("재접촉 관리"))

    return d, issues


# ─────────────────────────────────────────────────────────
# Dry-run / Apply 실행
# ─────────────────────────────────────────────────────────

def import_sheet(wb, sheet_name, *, apply_changes=False, schema_hint=None, skip_backup=False):
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"시트 없음: {sheet_name}")
    ws = wb[sheet_name]

    # 행 1~5 추출 → 스키마 감지
    rows_top = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
    schema, header_idx = detect_schema(rows_top)
    if schema_hint:
        schema = schema_hint
    if schema is None:
        raise SystemExit(f"스키마 감지 실패: {sheet_name}")

    # outlier: 25.3 / 25.6 처럼 row 2에 필드 헤더가 온 경우
    header_row = rows_top[header_idx - 1]
    headers_idx = header_index_map(header_row)
    data_start_row = header_idx + 1  # 헤더 다음 행이 데이터 시작

    report = {
        "sheet": sheet_name,
        "schema": schema,
        "header_row": header_idx,
        "rows_total": 0,
        "rows_imported": 0,
        "rows_skipped": 0,
        "issues": Counter(),
        "issue_samples": defaultdict(list),
        "unmapped_values": defaultdict(Counter),
        "patients_new": 0,
        "patients_matched": 0,
        "duplicate_candidates": [],
    }

    if schema == "A":
        raise SystemExit(f"Schema A(구식)는 별도 도구 필요. 이 시트: {sheet_name}")
    # Schema B/C는 동일 파서로 처리 (parse_schema_c_row가 header 이름 기반이라 호환)

    rows_to_apply = []

    name_idx = headers_idx.get("환자이름")
    date_idx = headers_idx.get("상담일자")
    for ri, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True),
                             start=data_start_row):
        # 빈 행: 환자이름·상담일자 둘 다 비어있으면 스킵 (행번호 컬럼만 채워진 빈 행 제거)
        nm = row[name_idx] if name_idx is not None and name_idx < len(row) else None
        dt = row[date_idx] if date_idx is not None and date_idx < len(row) else None
        if (nm is None or str(nm).strip() == "") and (dt is None or str(dt).strip() == ""):
            continue
        report["rows_total"] += 1
        try:
            parsed, issues = parse_schema_c_row(headers_idx, row)
        except Exception as e:
            report["rows_skipped"] += 1
            report["issues"][f"parse-error: {type(e).__name__}"] += 1
            continue

        if parsed is None:
            report["rows_skipped"] += 1
            for it in issues:
                report["issues"][it] += 1
            continue

        # NOT NULL 필수 — consult_date 없으면 적재 스킵
        if not parsed.get("consult_date"):
            report["rows_skipped"] += 1
            report["issues"]["consult_date 없음 — 적재 불가"] += 1
            for it in issues:
                report["issues"][it] += 1
            continue

        # Issue 기록
        for it in issues:
            report["issues"][it] += 1
            if len(report["issue_samples"][it]) < 5:
                report["issue_samples"][it].append({
                    "row": ri, "patient": parsed.get("patient_name"),
                })

        # 필드별 값 분포 추적 (리스트는 문자열로 평탄화)
        for fld in ("current_location_type", "referral_source_type",
                    "referral_source_detail", "consult_channel",
                    "admission_status", "admission_purpose"):
            val = parsed.get(fld)
            if not val:
                continue
            if isinstance(val, list):
                for v in val:
                    report["unmapped_values"][fld][str(v)] += 1
            else:
                report["unmapped_values"][fld][str(val)] += 1

        rows_to_apply.append((ri, parsed))

    # ── 적재 (apply 모드) ──
    if apply_changes:
        if not skip_backup:
            backup_db()
        conn = models.get_db()
        try:
            for ri, parsed in rows_to_apply:
                pid = _upsert_patient(conn, parsed, report)
                _insert_consultation(conn, pid, parsed)
                report["rows_imported"] += 1
            conn.commit()
        finally:
            conn.close()
    else:
        # dry-run: 환자 매칭 후보만 시뮬
        conn = models.get_db()
        try:
            for ri, parsed in rows_to_apply:
                _simulate_patient_match(conn, parsed, report)
        finally:
            conn.close()
        report["rows_imported"] = len(rows_to_apply)

    return report


def _simulate_patient_match(conn, parsed, report):
    name = parsed["patient_name"]
    phone = parsed.get("guardian_phone")
    row = None
    if phone:
        row = conn.execute(
            "SELECT id, name, guardian_phone FROM patients WHERE name=? AND guardian_phone=? LIMIT 1",
            (name, phone),
        ).fetchone()
    if row:
        report["patients_matched"] += 1
    else:
        # 같은 이름 다른 전화 = 중복 후보
        if phone:
            other = conn.execute(
                "SELECT id, guardian_phone FROM patients WHERE name=? AND guardian_phone IS NOT NULL AND guardian_phone <> ?",
                (name, phone),
            ).fetchall()
            if other:
                report["duplicate_candidates"].append({
                    "name": name, "new_phone": phone,
                    "existing_phones": [r["guardian_phone"] for r in other],
                })
        report["patients_new"] += 1


def _upsert_patient(conn, parsed, report):
    """find_or_create_patient의 인라인 버전 (transaction 한 번에 묶기 위해)."""
    name = parsed["patient_name"]
    phone = parsed.get("guardian_phone")
    row = None
    if phone:
        row = conn.execute(
            "SELECT id FROM patients WHERE name=? AND guardian_phone=? LIMIT 1",
            (name, phone),
        ).fetchone()
    if not row:
        row = conn.execute(
            "SELECT id FROM patients WHERE name=? AND (guardian_phone IS NULL OR guardian_phone='') LIMIT 1",
            (name,),
        ).fetchone()
    p_fields = {
        "gender": parsed.get("gender"),
        "residence_sido": parsed.get("residence_sido"),
        "residence_sigungu": parsed.get("residence_sigungu"),
        "guardian_name": parsed.get("guardian_name"),
        "guardian_relation": parsed.get("guardian_relation"),
        "guardian_phone": phone,
    }
    if row:
        pid = row["id"]
        sets, vals = [], []
        for k, v in p_fields.items():
            if v not in (None, ""):
                sets.append(f"{k}=?"); vals.append(v)
        if sets:
            sets.append("updated_at=CURRENT_TIMESTAMP")
            vals.append(pid)
            conn.execute(f"UPDATE patients SET {', '.join(sets)} WHERE id=?", vals)
        report["patients_matched"] += 1
        return pid
    cur = conn.execute(
        """INSERT INTO patients (name, gender, residence_sido, residence_sigungu,
                                  guardian_name, guardian_relation, guardian_phone)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (name, p_fields["gender"], p_fields["residence_sido"], p_fields["residence_sigungu"],
         p_fields["guardian_name"], p_fields["guardian_relation"], phone),
    )
    report["patients_new"] += 1
    return cur.lastrowid


def _insert_consultation(conn, pid, parsed):
    fields = {
        "consult_date": parsed.get("consult_date"),
        "counselor": parsed.get("counselor"),
        "consult_channel": parsed.get("consult_channel"),
        "patient_age": parsed.get("patient_age"),
        "current_location_type": parsed.get("current_location_type"),
        "current_location_name": parsed.get("current_location_name"),
        "source_hospital": parsed.get("source_hospital"),
        "admission_purpose": parsed.get("admission_purpose"),
        "disease_detail": parsed.get("disease_detail"),
        "admission_status": parsed.get("admission_status"),
        "actual_admission_date": parsed.get("actual_admission_date"),
        "recontact_memo": parsed.get("recontact_memo"),
        "referrer_person": parsed.get("referrer_person"),
        "import_source": "excel",
    }
    # JSON 필드
    rt = parsed.get("referral_source_type")
    if rt:
        fields["referral_source_type"] = json.dumps(rt, ensure_ascii=False)
    rd = parsed.get("referral_source_detail")
    if rd:
        fields["referral_source_detail"] = json.dumps(rd, ensure_ascii=False)
    dx = parsed.get("diseases")
    if dx:
        fields["diseases"] = json.dumps(dx, ensure_ascii=False)

    cols = [k for k, v in fields.items() if v not in (None, "")]
    vals = [fields[k] for k in cols]
    conn.execute(
        f"INSERT INTO consultations (patient_id, {','.join(cols)}) VALUES ({','.join(['?']*(len(cols)+1))})",
        [pid] + vals,
    )


def backup_db():
    src = ROOT / "bokju.db"
    dst = ROOT / "backups" / f"pre_excel_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    dst.parent.mkdir(exist_ok=True)
    shutil.copy2(src, dst)
    print(f"[backup] {dst}")


# ─────────────────────────────────────────────────────────
# 리포트 출력
# ─────────────────────────────────────────────────────────

def render_report(report, apply_mode=False):
    """리포트를 문자열로 직렬화."""
    lines = []
    title = "적재 완료" if apply_mode else "Dry-Run 리포트"
    lines.append(f"=== {title} : {report['sheet']} (Schema {report['schema']}) ===")
    lines.append(f"  헤더 위치: 행 {report['header_row']}")
    lines.append(f"  데이터 행: {report['rows_total']}")
    lines.append(f"  성공:      {report['rows_imported']}")
    lines.append(f"  스킵:      {report['rows_skipped']}")
    lines.append(f"  환자 신규: {report['patients_new']}")
    lines.append(f"  환자 매칭: {report['patients_matched']}")

    if report["issues"]:
        lines.append("")
        lines.append("  [이슈 종류 / 발생 횟수]")
        for issue, n in sorted(report["issues"].items(), key=lambda x: -x[1]):
            lines.append(f"    {n:4d} × {issue}")
            for sample in report["issue_samples"][issue][:3]:
                lines.append(f"           ↳ row {sample['row']}: {sample['patient']}")

    if report["unmapped_values"]:
        lines.append("")
        lines.append("  [필드별 distinct 값 분포]")
        for fld, counter in report["unmapped_values"].items():
            lines.append(f"    {fld}:")
            for val, n in counter.most_common(15):
                lines.append(f"      {n:4d} × {val}")

    if report["duplicate_candidates"]:
        lines.append("")
        lines.append(f"  [환자 중복 후보 ({len(report['duplicate_candidates'])}건)]")
        for c in report["duplicate_candidates"][:10]:
            lines.append(f"    {c['name']}: 신규 전화 {c['new_phone']} vs 기존 {c['existing_phones']}")
    return "\n".join(lines) + "\n"


# ─────────────────────────────────────────────────────────
# 진입점
# ─────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path", help="엑셀 파일 경로")
    ap.add_argument("--sheet", help="대상 시트 이름 (단일)")
    ap.add_argument("--sheets", help="대상 시트 콤마 구분 (다중)")
    ap.add_argument("--apply", action="store_true", help="실제 DB에 적재")
    ap.add_argument("--schema", help="스키마 강제 지정 (A/B/C)")
    ap.add_argument("--report-file", help="리포트를 UTF-8 파일로 저장")
    args = ap.parse_args()

    xlsx = Path(args.xlsx_path)
    if not xlsx.exists():
        raise SystemExit(f"파일 없음: {xlsx}")

    if not args.sheet and not args.sheets:
        raise SystemExit("--sheet 또는 --sheets 중 하나를 지정")
    sheet_list = [args.sheet] if args.sheet else [s.strip() for s in args.sheets.split(",")]

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    if args.apply and len(sheet_list) > 1:
        backup_db()  # 다중 시트 적재 시 백업 한 번만

    parts = []
    for s in sheet_list:
        rep = import_sheet(wb, s, apply_changes=args.apply,
                           schema_hint=args.schema, skip_backup=True)
        parts.append(render_report(rep, apply_mode=args.apply))
    text = "\n".join(parts)
    if args.report_file:
        Path(args.report_file).write_text(text, encoding="utf-8")
        print(f"[report] {args.report_file}")
    else:
        try:
            print(text)
        except UnicodeEncodeError:
            sys.stdout.buffer.write(text.encode("utf-8"))


if __name__ == "__main__":
    main()
