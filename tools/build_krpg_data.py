"""KRPG 2.2 엑셀의 세 시트를 웹 검색용 JSON으로 변환한다.

사용법:
  python tools/build_krpg_data.py 원본.xlsx [출력.json]
"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


SHEETS = {
    "business": ("사업대상(1,477개)", 4, 1477),
    "changes": ("KCD 변경 목록(570개)", 3, 570),
    "all": ("KRPG전체(4,585개)", 3, 4585),
}


def clean(value):
    return str(value).strip() if value is not None else ""


def extract_sheet(wb, sheet_name, first_row, expected_count):
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"시트를 찾을 수 없습니다: {sheet_name}")
    rows = []
    for values in wb[sheet_name].iter_rows(min_row=first_row, values_only=True):
        seq, kric, kcd, name_ko, name_en, note = values[:6]
        if not clean(kcd):
            continue
        rows.append({
            "seq": int(seq) if str(seq).isdigit() else clean(seq),
            "kric": clean(kric).zfill(2),
            "kcd": clean(kcd).upper(),
            "name_ko": clean(name_ko),
            "name_en": clean(name_en),
            "note": clean(note),
        })
    if len(rows) != expected_count:
        raise SystemExit(f"{sheet_name} 행 수 불일치: {len(rows):,}개 (예상 {expected_count:,}개)")
    return rows


def main():
    if len(sys.argv) < 2:
        raise SystemExit("원본 xlsx 경로를 입력하세요.")
    source = Path(sys.argv[1])
    output = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("data/krpg_v22.json")
    wb = load_workbook(source, read_only=True, data_only=True)
    datasets = {
        key: extract_sheet(wb, sheet_name, first_row, expected)
        for key, (sheet_name, first_row, expected) in SHEETS.items()
    }
    payload = {
        "title": "한국형 재활환자분류체계(KRPG) 버전 2.2 KCD 진단 코드 목록",
        "version": "2.2",
        "counts": {key: len(rows) for key, rows in datasets.items()},
        "source_file": source.name,
        "datasets": datasets,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"{output}: " + ", ".join(f"{key} {len(rows):,}개" for key, rows in datasets.items()))


if __name__ == "__main__":
    main()
