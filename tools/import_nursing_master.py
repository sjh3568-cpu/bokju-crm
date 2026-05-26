"""전국 요양원(노인의료복지시설) 공식명 마스터 import.

데이터 출처 (사용자가 직접 다운로드):
- 보건복지부 노인의료복지시설 현황 (대표: e-나라지표·복지부 발표 자료)
- 국민건강보험공단 노인장기요양보험 공시 자료
- 공공데이터포털(data.go.kr) — "노인복지시설 현황" 키워드

사용 예:
  python tools/import_nursing_master.py --xlsx 요양원_현황.xlsx
  python tools/import_nursing_master.py --csv nursing_homes.csv

CSV/XLSX 헤더는 다음 후보 중 자동 인식: 시설명/기관명/요양원명, 시도/지역/소재지,
종별/시설종류/구분, 주소/소재지주소, 전화/대표전화, 코드/시설코드/요양기관기호.
"""
from __future__ import annotations

import argparse
import csv
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import models  # noqa: E402


# 풀 시도명 → source_hospitals 표기와 동일한 짧은 명칭. region 일관성 보장.
SIDO_SHORT = {
    "서울특별시": "서울", "부산광역시": "부산", "대구광역시": "대구",
    "인천광역시": "인천", "광주광역시": "광주", "대전광역시": "대전",
    "울산광역시": "울산", "세종특별자치시": "세종시",
    "경기도": "경기", "강원특별자치도": "강원", "강원도": "강원",
    "충청북도": "충북", "충청남도": "충남",
    "전라북도": "전북", "전북특별자치도": "전북",
    "전라남도": "전남",
    "경상북도": "경북", "경상남도": "경남",
    "제주특별자치도": "제주",
}


def _normalize_sido(region: str) -> str:
    """풀명·약명 모두 SOURCE_HOSPITAL_SEED와 동일한 짧은 표기로 정규화."""
    if not region:
        return ""
    r = region.strip()
    return SIDO_SHORT.get(r, r)  # 이미 짧은 명칭이면 그대로


def _pick(row: dict, *keys: str) -> str:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return str(row[key]).strip()
    lower = {str(k).strip().lower(): v for k, v in row.items()}
    for key in keys:
        val = lower.get(key.lower())
        if val not in (None, ""):
            return str(val).strip()
    return ""


def _entry_from_row(row: dict) -> dict:
    addr = _pick(row, "주소", "소재지주소", "소재지", "기관별 상세주소", "address")
    region = _pick(row, "시도", "시도명", "지역", "sido")
    # NHIS 양식: '시도 시군구 법정동명' 한 칸에 풀명. 첫 토큰을 region(시도 약명)으로 사용.
    if not region:
        fullname = _pick(row, "시도 시군구 법정동명")
        if fullname:
            region = fullname.split()[0] if fullname.split() else ""
    if not region and addr:
        region = addr.split()[0] if addr.split() else ""
    return {
        "name": _pick(row, "시설명", "기관명", "요양원명",
                      "장기요양기관명", "장기요양기관이름", "name"),
        "region": _normalize_sido(region),
        "kind": _pick(row, "종별", "시설종류", "시설구분", "구분",
                      "급여종류명", "급여종류", "kind"),
        "address": addr,
        "phone": _pick(row, "전화번호", "대표전화", "연락처", "phone"),
        "official_code": _pick(row, "시설코드", "요양기관기호",
                               "장기요양기관번호", "장기요양기관코드", "code"),
    }


def read_xlsx(path: Path) -> list[dict]:
    import openpyxl  # 지연 import
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    # 헤더 행 탐색 — 첫 비어있지 않은 행이 헤더라고 가정. '시설명/기관명' 키워드로 검증.
    headers: list[str] = []
    for row in iterator:
        if not row:
            continue
        candidate = [str(c).strip() if c is not None else "" for c in row]
        if any(k in candidate for k in ("시설명", "기관명", "요양원명", "장기요양기관명")):
            headers = candidate
            break
        # 첫 행이 헤더가 아니면 다음으로
    if not headers:
        # fallback: 첫 행을 헤더로
        wb.close()
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        iterator = ws.iter_rows(values_only=True)
        headers = [str(c).strip() if c is not None else "" for c in next(iterator)]
    entries: list[dict] = []
    for row in iterator:
        if not row or not any(row):
            continue
        rec = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        entry = _entry_from_row(rec)
        if entry.get("name"):
            entries.append(entry)
    return entries


def read_csv(path: Path, *, encoding: str) -> list[dict]:
    with path.open("r", encoding=encoding, newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t|")
        rows = csv.DictReader(f, dialect=dialect)
        return [entry for row in rows if (entry := _entry_from_row(row)).get("name")]


def main() -> int:
    p = argparse.ArgumentParser(description="요양원(노인의료복지시설) 마스터 import")
    src = p.add_mutually_exclusive_group(required=True)
    src.add_argument("--xlsx", type=Path, help="보건복지부/공단 xlsx 경로")
    src.add_argument("--csv", type=Path, help="요양원 목록 CSV/TSV 경로")
    p.add_argument("--encoding", default="utf-8-sig")
    args = p.parse_args()

    models.init_db()
    if args.xlsx:
        entries = read_xlsx(args.xlsx)
        source = "mohw_xlsx"
    else:
        entries = read_csv(args.csv, encoding=args.encoding)
        source = "mohw_csv"

    imported = models.upsert_source_nursing_homes(entries, source=source)
    print(f"imported={imported} source={source}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
